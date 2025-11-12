"""
Time Use Survey 2022 - Time Estimate Application
This application allows users to select demographic attributes and get
detailed estimates of average time spent on activities with bootstrap variance estimates.
"""

import streamlit as st
import pandas as pd
import numpy as np
import pyreadstat
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Set page config
st.set_page_config(
    page_title="Time Use Survey 2022 - Time Estimates",
    page_icon="⏱️",
    layout="wide"
)

# Data paths
DATA_DIR = Path("TU_ET_2022/Data_Données")
MAIN_FILE = DATA_DIR / "TU_ET_2022_Main-Principal_PUMF.sas7bdat"

# Value label mappings for filter variables
VALUE_LABELS = {
    'GENDER2': {
        1: "Men+",
        2: "Women+",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'PRV': {
        10: "Newfoundland and Labrador",
        11: "Prince Edward Island",
        12: "Nova Scotia",
        13: "New Brunswick",
        24: "Quebec",
        35: "Ontario",
        46: "Manitoba",
        47: "Saskatchewan",
        48: "Alberta",
        59: "British Columbia",
        96: "Valid skip",
        97: "Don't know",
        98: "Refusal",
        99: "Not stated"
    },
    'REGION': {
        1: "Atlantic region",
        2: "Quebec",
        3: "Ontario",
        4: "Prairie region",
        5: "British Columbia",
        6: "Territories",
        "01": "Atlantic region",
        "02": "Quebec",
        "03": "Ontario",
        "04": "Prairie region",
        "05": "British Columbia",
        "06": "Territories",
        96: "Valid skip",
        97: "Don't know",
        98: "Refusal",
        99: "Not stated"
    },
    'AGEGR10': {
        1: "15 to 24 years",
        2: "25 to 34 years",
        3: "35 to 44 years",
        4: "45 to 54 years",
        5: "55 to 64 years",
        6: "65 to 74 years",
        7: "75 years and over",
        96: "Valid skip",
        97: "Don't know",
        98: "Refusal",
        99: "Not stated"
    },
    'MARSTAT': {
        1: "Married",
        2: "Living common-law",
        3: "Never married (not living common law)",
        4: "Separated (not living common law)",
        5: "Divorced (not living common law)",
        6: "Widowed (not living common law)",
        96: "Valid skip",
        97: "Don't know",
        98: "Refusal",
        99: "Not stated"
    },
    'CHH0017C': {
        0: "No children",
        1: "One child",
        2: "Two children",
        3: "Three children or more",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'CHH0004C': {
        0: "No children",
        1: "One or more children",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'CHH0514C': {
        0: "No children",
        1: "One or more children",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'CHH1517C': {
        0: "No children",
        1: "One or more children",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'PHSDFLG': {
        1: "Yes",
        2: "No",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'ACT7DAYC': {
        1: "Working at a paid job or business...",
        2: "Going to school",
        3: "Household work /caring for children",
        4: "Retired",
        5: "Other...",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'MRW_D40A': {
        1: "Yes",
        2: "No",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'MRW_D40B': {
        1: "Yes",
        2: "No",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'MAP_110C': {
        1: "Working at a paid job or business...",
        2: "Household work /caring for children",
        3: "Retired",
        4: "Other...",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'MAP_D40A': {
        1: "Yes",
        2: "No",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'ED_05': {
        1: "Less than high school diploma or its equivalent",
        2: "High school diploma or a high school equivalency certificate",
        3: "Trades certificate or diploma",
        4: "College, CEGEP or other non-university certificate or diploma",
        5: "University certificate or diploma below the bachelor's level",
        6: "Bachelor's degree",
        7: "University certificate, diploma, or degree above the bachelor's level",
        96: "Valid skip",
        97: "Don't know",
        98: "Refusal",
        99: "Not stated"
    },
    'INC_C': {
        1: "Less than $50,000",
        2: "$50,000 to $74,999",
        3: "$75,000 to $99,999",
        4: "$100,000 to $124,999",
        5: "$125,000 and over",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'FAMINC_C': {
        1: "Less than $50,000",
        2: "$50,000 to $74,999",
        3: "$75,000 to $99,999",
        4: "$100,000 to $124,999",
        5: "$125,000 and over",
        6: "Valid skip",
        7: "Don't know",
        8: "Refusal",
        9: "Not stated"
    },
    'AGEPRC': {
        1: "15 to 34 years",
        2: "35 to 44 years",
        3: "45 to 54 years",
        4: "55 to 64 years",
        5: "65 to 74 years",
        6: "75 years and over",
        96: "Valid skip",
        97: "Don't know",
        98: "Refusal",
        99: "Not stated"
    }
}

def format_value(var_name, value):
    """Format a value using its label if available"""
    if var_name in VALUE_LABELS:
        # Try exact match first
        if value in VALUE_LABELS[var_name]:
            return VALUE_LABELS[var_name][value]
        # Try converting to int/string if needed
        try:
            if isinstance(value, str) and value.isdigit():
                int_val = int(value)
                if int_val in VALUE_LABELS[var_name]:
                    return VALUE_LABELS[var_name][int_val]
        except:
            pass
        try:
            if isinstance(value, (int, float)):
                str_val = f"{int(value):02d}" if var_name == 'REGION' else str(value)
                if str_val in VALUE_LABELS[var_name]:
                    return VALUE_LABELS[var_name][str_val]
        except:
            pass
    return str(value)

# Activity category mappings
ACTIVITY_CATEGORIES = {
    "Sleep and Rest": ["DUR101", "DUR102", "DUR103", "DUR104", "DUR109"],
    "Personal Care": ["DUR126", "DUR127", "DUR128", "DUR129", "DUR130", "DUR199"],
    "Eating and Drinking": ["DUR151", "DUR152", "DUR153", "DUR154", "DUR159"],
    "Household Work": [
        "DUR201", "DUR202", "DUR203", "DUR204", "DUR205", "DUR206", "DUR207", 
        "DUR208", "DUR209", "DUR231", "DUR232", "DUR233", "DUR234", "DUR235",
        "DUR236", "DUR237", "DUR238", "DUR239", "DUR240", "DUR241", "DUR299"
    ],
    "Shopping": ["DUR261", "DUR262", "DUR263", "DUR264", "DUR269"],
    "Childcare": [
        "DUR301", "DUR302", "DUR303", "DUR304", "DUR305", "DUR306", "DUR307", "DUR399"
    ],
    "Adult Care": ["DUR351", "DUR352", "DUR353", "DUR359"],
    "Travel": [
        "DUR401", "DUR402", "DUR403", "DUR404", "DUR405", "DUR406", "DUR407",
        "DUR408", "DUR409", "DUR410", "DUR411", "DUR412", "DUR413", "DUR414",
        "DUR415", "DUR416", "DUR499"
    ],
    "Paid Work": [
        "DUR501", "DUR502", "DUR503", "DUR504", "DUR505", "DUR506", "DUR599"
    ],
    "Education": ["DUR601", "DUR602", "DUR603", "DUR604", "DUR699"],
    "Socializing": ["DUR701", "DUR702", "DUR799"],
    "Volunteering and Unpaid Work": [
        "DUR801", "DUR802", "DUR803", "DUR804", "DUR805", "DUR806", "DUR807", 
        "DUR808", "DUR899"
    ],
    "Civic and Religious": ["DUR901", "DUR902", "DUR903", "DUR999"],
    "Sports and Exercise": [
        "DUR1001", "DUR1002", "DUR1004", "DUR1005", "DUR1099"
    ],
    "Culture and Leisure": [
        "DUR1101", "DUR1102", "DUR1103", "DUR1104", "DUR1105", "DUR1106", "DUR1199"
    ],
    "Mass Media": ["DUR1201", "DUR1202", "DUR1203", "DUR1204", "DUR1299"],
    "Other Activities": ["DUR1301", "DUR1302", "DUR1303", "DUR1304", "DUR9999"]
}

# Activity descriptions (from SAS labels)
ACTIVITY_DESCRIPTIONS = {
    "DUR101": "Essential sleep (night or day)",
    "DUR102": "Sleeplessness, insomnia",
    "DUR103": "Naps, lying down, resting, relaxing",
    "DUR104": "Sick in bed, bed or rehabilitative rest, convalescence",
    "DUR109": "Sleep/relaxing/bed rest, unspecified",
    "DUR126": "Personal care",
    "DUR127": "Self-administered medical care",
    "DUR128": "Health professional visit, consultation",
    "DUR129": "Receiving personal care from another household member",
    "DUR130": "Receiving personal care from other personal care providers",
    "DUR151": "Eating",
    "DUR152": "Drinking other than with meals or snacks",
    "DUR153": "Break or lunch related to paid work activities",
    "DUR154": "Break or lunch related to studying or learning",
    "DUR159": "Eating or drinking, unspecified",
    "DUR199": "Own personal care, unspecified",
    "DUR201": "Preparing or serving meals or snacks",
    "DUR202": "Food (or meal) cleanup, dish washing",
    "DUR203": "Preserving foods",
    "DUR204": "Unpacking groceries",
    "DUR205": "Indoor house cleaning, tidying, care of house plants",
    "DUR206": "Taking out garbage, recycling, compost, unpacking of goods",
    "DUR207": "Laundry, mending, ironing, shoe care, etc.",
    "DUR208": "Organizing, planning, paying bills, managing mail",
    "DUR209": "Pet care",
    "DUR231": "Sewing clothes (for self or hhld. or fam.)",
    "DUR232": "Interior do-it-yourself improvement, maintenance",
    "DUR233": "Installation, servicing or repair of household goods",
    "DUR234": "Packing or unpacking for a trip or camping",
    "DUR235": "Packing or unpacking for a move of the household",
    "DUR236": "Outdoor cleaning",
    "DUR237": "Exterior D.I.Y. improvement, maintenance, home repair",
    "DUR238": "Do-it-yourself construction",
    "DUR239": "Vehicle maintenance or repairs",
    "DUR240": "Harvesting, stacking or cutting firewood",
    "DUR241": "Gardening, raising animals, etc., for hhld. use",
    "DUR261": "In-person shopping for goods",
    "DUR262": "In-person shopping for services",
    "DUR263": "Online shopping for goods or services",
    "DUR264": "Researching for purchasing goods or services",
    "DUR269": "Shopping, unspecified",
    "DUR299": "Unpaid household work (hhld. or family), unspecified",
    "DUR301": "Childcare (<15) (hhld. or fam.): Personal care",
    "DUR302": "Childcare (<15) (hhld. or fam.): Read, play, talk",
    "DUR303": "Childcare (<15) (hhld. or fam.): Educational",
    "DUR304": "Childcare (<15) (hhld. or fam.): Accompanying",
    "DUR305": "Childcare (15-17) (hhld. or fam.): Personal care",
    "DUR306": "Childcare (15-17) (hhld. or fam.): Educational",
    "DUR307": "Childcare (15-17) (hhld. or fam.): Accompanying",
    "DUR351": "Adult care (18+) (hhld. or fam.): Personal care",
    "DUR352": "Adult care (18+) (hhld. or fam.): Accompanying",
    "DUR353": "Adult care (18+) (hhld. or fam.): Household chores",
    "DUR359": "Care of adults (18+) (hhld. or family), unspecified",
    "DUR399": "Care of children (<18) (hhld. or family), unspecified",
    "DUR401": "Travel to or from receiving personal care",
    "DUR402": "Travel related to caring for children (<18) (hhld. or fam.)",
    "DUR403": "Travel related to caring for adults (18+) (hhld. or fam.)",
    "DUR404": "Travel to or from eating or drinking",
    "DUR405": "Travel related to regular household tasks",
    "DUR406": "Travel related to occasional household tasks",
    "DUR407": "Travel to or from paid work activities",
    "DUR408": "Travel to or from studying or learning activities",
    "DUR409": "Travel to or from shopping",
    "DUR410": "Travel to or from socializing or communicating activities",
    "DUR411": "Travel to or from informal or org.-based volunteering",
    "DUR412": "Travel to or from civic/religious/community activities",
    "DUR413": "Travel to or from sports participation or physical exercise",
    "DUR414": "Travel to or from culture or sports events",
    "DUR415": "Travel to or from hobbies, leisure, outdoor activities",
    "DUR416": "Travel related to mass media activities",
    "DUR499": "Travel, unspecified",
    "DUR501": "Paid work",
    "DUR502": "Paid training",
    "DUR503": "Waiting or idle time related to paid work activities",
    "DUR504": "Selling of goods or services for pay or profit",
    "DUR505": "Other income-generating activities",
    "DUR506": "Looking for work",
    "DUR599": "Paid work activities, unspecified",
    "DUR601": "Schooling full time or part time - on site",
    "DUR602": "Schooling full time or part time - online",
    "DUR603": "Homework, studying or being tutored",
    "DUR604": "Self development, leisure or special interest classes",
    "DUR699": "Studying or learning, unspecified",
    "DUR701": "Socializing or communicating - in person",
    "DUR702": "Socializing or communicating - using any type of technology",
    "DUR799": "Socializing or communicating, unspecified",
    "DUR801": "Unpaid help (other hhlds, non-fam) - childcare (<18)",
    "DUR802": "Unpaid help (other hhlds, non-fam) - caring for an adult",
    "DUR803": "Unpaid help (other hhlds, non-fam) - doing chores",
    "DUR804": "Unpaid work in enterprises owned by other households",
    "DUR805": "Unpaid coaching or administering sports",
    "DUR806": "Organization-based volunteering (unpaid, non-compulsory)",
    "DUR807": "Unpaid work aimed at improving the community",
    "DUR808": "Unpaid work required by school, employer, court or other",
    "DUR899": "Unpaid work not for own household, unspecified",
    "DUR901": "Participating in community, cultural or social events",
    "DUR902": "Civic participation (voting, jury duty)",
    "DUR903": "Religious practices",
    "DUR999": "Civic, religious or community activities, unspecified",
    "DUR1001": "Exercising",
    "DUR1002": "Organized recreational sports",
    "DUR1004": "Outdoor sports (non-competitive)",
    "DUR1005": "Other sports activities",
    "DUR1099": "Sports participation and physical exercise, unspecified",
    "DUR1101": "Attending cinema",
    "DUR1102": "Attending a concert or live entertainment event",
    "DUR1103": "Attending sporting events",
    "DUR1104": "Museums, galleries, zoos, observatories, amusement parks",
    "DUR1105": "Arts, hobbies or playing games",
    "DUR1106": "Leisure or outdoor activities",
    "DUR1199": "Culture, hobbies, leisure, outdoor activities, unspecified",
    "DUR1201": "Reading",
    "DUR1202": "Watching television",
    "DUR1203": "Listening to music, radio or podcasts",
    "DUR1204": "Use of technology",
    "DUR1299": "Mass media activities, unspecified",
    "DUR1301": "Waiting time",
    "DUR1302": "Free time, thinking, smoking",
    "DUR1303": "Doing nothing",
    "DUR1304": "Other activities",
    "DUR9999": "Activity unspecified"
}

# Get all DUR variables
ALL_DUR_VARS = []
for category, vars_list in ACTIVITY_CATEGORIES.items():
    ALL_DUR_VARS.extend(vars_list)
ALL_DUR_VARS = sorted(set(ALL_DUR_VARS))

@st.cache_data
def load_data():
    """Load the main dataset"""
    try:
        # Try reading with pyreadstat first
        df, meta = pyreadstat.read_sas7bdat(str(MAIN_FILE))
        return df, meta
    except Exception as e1:
        try:
            # Fallback: try reading the text file if SAS file doesn't work
            txt_file = DATA_DIR / "TU_ET_2022_Main-Principal_PUMF.txt"
            if txt_file.exists():
                # This is a fixed-width file, we'll need to handle it differently
                st.warning("SAS file not readable, attempting to use text file. This may take longer.")
                # For now, return error - we'll need to implement text file reading
                st.error(f"Cannot read SAS file. Error: {e1}")
                return None, None
            else:
                st.error(f"Error loading data: {e1}")
                return None, None
        except Exception as e2:
            st.error(f"Error loading data: {e1}. Also tried text file: {e2}")
            return None, None

def get_bootstrap_weights(df):
    """Get all bootstrap weight column names"""
    # Main file uses WTBS_ weights (person-level bootstrap weights)
    wtbs_cols = [col for col in df.columns if col.startswith('WTBS_')]
    # Sort numerically by the number after WTBS_
    def sort_key(col):
        try:
            return int(col.split('_')[1])
        except:
            return 0
    return sorted(wtbs_cols, key=sort_key)

def calculate_weighted_mean(df, var, weight_col='WGHT_PER'):
    """Calculate weighted mean for a variable"""
    # Filter out missing values
    mask = df[var].notna() & (df[weight_col] > 0)
    if mask.sum() == 0:
        return np.nan
    
    weighted_sum = (df.loc[mask, var] * df.loc[mask, weight_col]).sum()
    total_weight = df.loc[mask, weight_col].sum()
    
    if total_weight == 0:
        return np.nan
    
    return weighted_sum / total_weight

def calculate_bootstrap_variance(df, var, weight_col='WGHT_PER', bootstrap_cols=None):
    """Calculate bootstrap variance using bootstrap weights"""
    if bootstrap_cols is None:
        bootstrap_cols = get_bootstrap_weights(df)
    
    if len(bootstrap_cols) == 0:
        return np.nan
    
    # Calculate estimate with main weight
    main_estimate = calculate_weighted_mean(df, var, weight_col)
    
    if np.isnan(main_estimate):
        return np.nan
    
    # Calculate estimates with each bootstrap weight
    bootstrap_estimates = []
    for bs_col in bootstrap_cols:
        if bs_col in df.columns:
            bs_estimate = calculate_weighted_mean(df, var, bs_col)
            if not np.isnan(bs_estimate):
                bootstrap_estimates.append(bs_estimate)
    
    if len(bootstrap_estimates) == 0:
        return np.nan
    
    # Bootstrap variance formula: sum((estimate_b - estimate_full)^2) / B
    bootstrap_estimates = np.array(bootstrap_estimates)
    variance = np.mean((bootstrap_estimates - main_estimate) ** 2)
    
    return variance

def filter_data(df, filters):
    """Apply filters to the dataset"""
    filtered_df = df.copy()
    
    for var, value in filters.items():
        if value is not None and var in filtered_df.columns:
            if isinstance(value, list):
                if len(value) > 0:
                    filtered_df = filtered_df[filtered_df[var].isin(value)]
            else:
                filtered_df = filtered_df[filtered_df[var] == value]
    
    return filtered_df

def get_unique_values(df, column):
    """Get unique non-null values from a column"""
    if column not in df.columns:
        return []
    unique_vals = df[column].dropna().unique()
    return sorted([v for v in unique_vals if pd.notna(v)])

def format_option_label(var_name, value):
    """Format option label for selectbox"""
    label = format_value(var_name, value)
    return f"{label} ({value})" if label != str(value) else str(value)

def main():
    st.title("⏱️ Time Use Survey 2022 - Time Estimates Application")
    st.markdown("""
    This application allows you to select demographic attributes and get detailed estimates 
    of average time spent on activities (in minutes per day) with bootstrap variance estimates.
    """)
    
    # Load data
    with st.spinner("Loading data..."):
        df, meta = load_data()
    
    if df is None:
        st.error("Failed to load data. Please check that the data files are in the correct location.")
        return
    
    st.success(f"Data loaded successfully! {len(df):,} records.")
    
    # Filters at the top of the page - reorganized as requested
    st.header("📊 Select Attributes")
    
    # Get unique values for filter variables
    filters = {}
    
    # Create columns for filters: Left, Middle, Right
    col1, col2, col3 = st.columns(3)
    
    # LEFT COLUMN: Geography, Demographics
    with col1:
        st.subheader("Geography")
        provinces = get_unique_values(df, 'PRV')
        if provinces:
            selected_provinces = st.multiselect(
                "Province of Residence",
                options=provinces,
                format_func=lambda x: format_option_label('PRV', x),
                help="Select one or more provinces. Leave empty to include all."
            )
            if len(selected_provinces) > 0:
                filters['PRV'] = selected_provinces
        
        regions = get_unique_values(df, 'REGION')
        if regions:
            selected_regions = st.multiselect(
                "Region of Residence",
                options=regions,
                format_func=lambda x: format_option_label('REGION', x),
                help="Select one or more regions. Leave empty to include all."
            )
            if len(selected_regions) > 0:
                filters['REGION'] = selected_regions
        
        st.subheader("Demographics")
        genders = get_unique_values(df, 'GENDER2')
        if genders:
            selected_genders = st.multiselect(
                "Gender",
                options=genders,
                format_func=lambda x: format_option_label('GENDER2', x),
                help="Select one or more gender categories. Leave empty to include all."
            )
            if len(selected_genders) > 0:
                filters['GENDER2'] = selected_genders
        
        age_groups = get_unique_values(df, 'AGEGR10')
        if age_groups:
            selected_ages = st.multiselect(
                "Age group (10-year groups)",
                options=age_groups,
                format_func=lambda x: format_option_label('AGEGR10', x),
                help="Select one or more age groups. Leave empty to include all."
            )
            if len(selected_ages) > 0:
                filters['AGEGR10'] = selected_ages
        
        marital_status = get_unique_values(df, 'MARSTAT')
        if marital_status:
            selected_marital = st.multiselect(
                "Marital status",
                options=marital_status,
                format_func=lambda x: format_option_label('MARSTAT', x),
                help="Select one or more marital statuses. Leave empty to include all."
            )
            if len(selected_marital) > 0:
                filters['MARSTAT'] = selected_marital
        
        education = get_unique_values(df, 'ED_05')
        if education:
            selected_edu = st.multiselect(
                "Educational attainment",
                options=education,
                format_func=lambda x: format_option_label('ED_05', x),
                help="Select one or more education levels. Leave empty to include all."
            )
            if len(selected_edu) > 0:
                filters['ED_05'] = selected_edu
    
    # MIDDLE COLUMN: Employment Status, Income
    with col2:
        st.subheader("Employment Status")
        employment_vars = {
            'MRW_D40A': 'Worked in the last 12 months',
            'MRW_D40B': 'Worked last week',
            'ACT7DAYC': 'Main activity - Last week'
        }
        
        for var, label in employment_vars.items():
            if var in df.columns:
                values = get_unique_values(df, var)
                if values:
                    selected = st.multiselect(
                        label,
                        options=values,
                        format_func=lambda x, v=var: format_option_label(v, x),
                        help="Select one or more options. Leave empty to include all."
                    )
                    if len(selected) > 0:
                        filters[var] = selected
        
        st.subheader("Income")
        income = get_unique_values(df, 'INC_C')
        if income:
            selected_inc = st.multiselect(
                "Income (grouped)",
                options=income,
                format_func=lambda x: format_option_label('INC_C', x),
                help="Select one or more income groups. Leave empty to include all."
            )
            if len(selected_inc) > 0:
                filters['INC_C'] = selected_inc
        
        fam_income = get_unique_values(df, 'FAMINC_C')
        if fam_income:
            selected_faminc = st.multiselect(
                "Family income (grouped)",
                options=fam_income,
                format_func=lambda x: format_option_label('FAMINC_C', x),
                help="Select one or more family income groups. Leave empty to include all."
            )
            if len(selected_faminc) > 0:
                filters['FAMINC_C'] = selected_faminc
    
    # RIGHT COLUMN: Spousal/Partner Information, Children
    with col3:
        st.subheader("Spousal/Partner Information")
        spouse_vars = {
            'PHSDFLG': 'Has spouse/partner in household',
            'MAP_110C': 'Main activity - Spouse/partner - 12 months',
            'MAP_D40A': 'Spouse/partner worked in last 12 months'
        }
        
        for var, label in spouse_vars.items():
            if var in df.columns:
                values = get_unique_values(df, var)
                if values:
                    selected = st.multiselect(
                        label,
                        options=values,
                        format_func=lambda x, v=var: format_option_label(v, x),
                        help="Select one or more options. Leave empty to include all."
                    )
                    if len(selected) > 0:
                        filters[var] = selected
        
        st.subheader("Children")
        num_children = get_unique_values(df, 'CHH0017C')
        if num_children:
            selected_children = st.multiselect(
                "Number of children (0-17 years)",
                options=num_children,
                format_func=lambda x: format_option_label('CHH0017C', x),
                help="Select one or more categories (e.g., select 1.0, 2.0, 3.0 to include all groups with children). Leave empty to include all."
            )
            if len(selected_children) > 0:
                filters['CHH0017C'] = selected_children
        
        child_age_vars = {
            'CHH0004C': 'Children 0-4 years (flag)',
            'CHH0514C': 'Children 5-14 years (flag)',
            'CHH1517C': 'Children 15-17 years (flag)'
        }
        
        for var, label in child_age_vars.items():
            if var in df.columns:
                values = get_unique_values(df, var)
                if values:
                    selected = st.multiselect(
                        label,
                        options=values,
                        format_func=lambda x, v=var: format_option_label(v, x),
                        help="Select one or more options. Leave empty to include all."
                    )
                    if len(selected) > 0:
                        filters[var] = selected
    
    # Update session state with current filters for real-time updates
    st.session_state.filters = filters
    
    # Calculate and display matching records count in real-time
    filtered_df = filter_data(df, filters)
    filtered_count = len(filtered_df)
    
    # Display matching records count box
    if filtered_count == 0:
        st.warning("**0 records** match your selected criteria. Please adjust your selections.")
    else:
        st.info(f"**{filtered_count:,} records** match your selected criteria.")
    
    # Store filtered count in session state
    st.session_state.filtered_count = filtered_count
    
    # Main content area
    st.header("📈 Time Estimates")
    
    if len(filtered_df) == 0:
        return
    
    # Calculate estimates
    if st.button("Calculate Estimates", type="primary"):
        bootstrap_cols = get_bootstrap_weights(filtered_df)
        if len(bootstrap_cols) == 0:
            st.error("No bootstrap weights found in the dataset. Cannot calculate variance estimates.")
            return
        
        st.info(f"Using {len(bootstrap_cols)} bootstrap weights for variance estimation.")
        
        with st.spinner("Calculating estimates and bootstrap variances (this may take a few minutes)..."):
            results = []
            
            # Get DUR variables that exist in the data
            available_dur_vars = [var for var in ALL_DUR_VARS if var in filtered_df.columns]
            
            if len(available_dur_vars) == 0:
                st.error("No DUR variables found in the dataset.")
                return
            
            # Calculate for each DUR variable
            progress_bar = st.progress(0)
            status_text = st.empty()
            total_vars = len(available_dur_vars)
            
            for idx, var in enumerate(available_dur_vars):
                status_text.text(f"Processing {var} ({idx + 1}/{total_vars})...")
                
                mean_est = calculate_weighted_mean(filtered_df, var)
                variance = calculate_bootstrap_variance(filtered_df, var, bootstrap_cols=bootstrap_cols)
                std_error = np.sqrt(variance) if not np.isnan(variance) else np.nan
                
                # Find category
                category = "Other"
                for cat, vars_list in ACTIVITY_CATEGORIES.items():
                    if var in vars_list:
                        category = cat
                        break
                
                # Get activity description
                activity_desc = ACTIVITY_DESCRIPTIONS.get(var, "Activity description not available")
                
                results.append({
                    'Activity_Code': var,
                    'Activity_Description': activity_desc,
                    'Activity_Category': category,
                    'Mean_Minutes_Per_Day': mean_est,
                    'Variance': variance,
                    'Standard_Error': std_error,
                    'Coefficient_of_Variation': (std_error / mean_est * 100) if not np.isnan(mean_est) and mean_est != 0 else np.nan
                })
                
                progress_bar.progress((idx + 1) / total_vars)
            
            st.session_state.results = pd.DataFrame(results)
            progress_bar.empty()
            status_text.empty()
            
            # Calculate category results immediately after individual results
            with st.spinner("Calculating aggregated category estimates..."):
                category_results_list = []
                bootstrap_cols = get_bootstrap_weights(filtered_df)
                
                for category, vars_list in ACTIVITY_CATEGORIES.items():
                    # Get variables in this category that exist in the data
                    cat_vars = [v for v in vars_list if v in filtered_df.columns]
                    
                    if len(cat_vars) == 0:
                        continue
                    
                    # Create a sum variable for this category
                    filtered_df['_CAT_SUM'] = filtered_df[cat_vars].sum(axis=1)
                    
                    # Calculate weighted mean for the sum
                    mean_est = calculate_weighted_mean(filtered_df, '_CAT_SUM')
                    
                    # Calculate bootstrap variance for the sum
                    variance = calculate_bootstrap_variance(filtered_df, '_CAT_SUM', bootstrap_cols=bootstrap_cols)
                    std_error = np.sqrt(variance) if not np.isnan(variance) else np.nan
                    
                    category_results_list.append({
                        'Activity_Category': category,
                        'Mean_Minutes_Per_Day': mean_est,
                        'Variance': variance,
                        'Standard_Error': std_error,
                        'Coefficient_of_Variation': (std_error / mean_est * 100) if not np.isnan(mean_est) and mean_est != 0 else np.nan
                    })
                
                st.session_state.category_results = pd.DataFrame(category_results_list)
            
            st.success("Calculations complete!")
    
    # Display results
    if 'results' in st.session_state and st.session_state.results is not None:
        results_df = st.session_state.results.copy()
        
        # Round numeric columns
        numeric_cols = ['Mean_Minutes_Per_Day', 'Variance', 'Standard_Error', 'Coefficient_of_Variation']
        for col in numeric_cols:
            if col in results_df.columns:
                results_df[col] = results_df[col].round(2)
        
        # Display by activity code
        st.subheader("By Activity Code")
        # Reorder columns to show description first
        display_cols = ['Activity_Code', 'Activity_Description', 'Activity_Category'] + [c for c in results_df.columns if c not in ['Activity_Code', 'Activity_Description', 'Activity_Category']]
        display_df = results_df[[c for c in display_cols if c in results_df.columns]]
        st.dataframe(display_df, use_container_width=True, height=400)
        
        # Display by category
        if 'category_results' in st.session_state and st.session_state.category_results is not None:
            st.subheader("By Activity Category")
            category_results = st.session_state.category_results.copy()
            
            # Round
            for col in numeric_cols:
                if col in category_results.columns:
                    category_results[col] = category_results[col].round(2)
            
            st.dataframe(category_results, use_container_width=True, height=400)
        
        # Export options - Single download button
        st.subheader("📥 Export Results")
        
        try:
            from io import BytesIO
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            output = BytesIO()
            
            # Get all available filter variables and their options
            all_filter_vars = {
                'PRV': get_unique_values(df, 'PRV'),
                'REGION': get_unique_values(df, 'REGION'),
                'GENDER2': get_unique_values(df, 'GENDER2'),
                'AGEGR10': get_unique_values(df, 'AGEGR10'),
                'MARSTAT': get_unique_values(df, 'MARSTAT'),
                'ED_05': get_unique_values(df, 'ED_05'),
                'MRW_D40A': get_unique_values(df, 'MRW_D40A'),
                'MRW_D40B': get_unique_values(df, 'MRW_D40B'),
                'ACT7DAYC': get_unique_values(df, 'ACT7DAYC'),
                'PHSDFLG': get_unique_values(df, 'PHSDFLG'),
                'MAP_110C': get_unique_values(df, 'MAP_110C'),
                'MAP_D40A': get_unique_values(df, 'MAP_D40A'),
                'CHH0017C': get_unique_values(df, 'CHH0017C'),
                'CHH0004C': get_unique_values(df, 'CHH0004C'),
                'CHH0514C': get_unique_values(df, 'CHH0514C'),
                'CHH1517C': get_unique_values(df, 'CHH1517C'),
                'INC_C': get_unique_values(df, 'INC_C'),
                'FAMINC_C': get_unique_values(df, 'FAMINC_C')
            }
            
            # Create single sheet with all sections
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Create empty dataframe to start
                all_data = []
                
                # TOP SECTION: Source and Filters
                all_data.append(["Time Use Survey 2022 - Time Estimates"])
                all_data.append([""])
                all_data.append(["Source:"])
                all_data.append(["Statistics Canada. General Social Survey - Time Use Component, 2022. " +
                                "Public Use Microdata File. Statistics Canada Catalogue no. 45M0001X. " +
                                "This does not constitute an endorsement by Statistics Canada of this product."])
                all_data.append([""])
                all_data.append(["Generated:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")])
                all_data.append([""])
                all_data.append(["Filter Criteria:"])
                all_data.append(["Variable", "Selected Value", "All Available Options"])
                
                # Add filter information
                filter_labels = {
                    'PRV': 'Province of Residence',
                    'REGION': 'Region of Residence',
                    'GENDER2': 'Gender',
                    'AGEGR10': 'Age Group (10-year groups)',
                    'MARSTAT': 'Marital Status',
                    'ED_05': 'Educational Attainment',
                    'MRW_D40A': 'Worked in the last 12 months',
                    'MRW_D40B': 'Worked last week',
                    'ACT7DAYC': 'Main activity - Last week',
                    'PHSDFLG': 'Has spouse/partner in household',
                    'MAP_110C': 'Main activity - Spouse/partner - 12 months',
                    'MAP_D40A': 'Spouse/partner worked in last 12 months',
                    'CHH0017C': 'Number of children (0-17 years)',
                    'CHH0004C': 'Children 0-4 years (flag)',
                    'CHH0514C': 'Children 5-14 years (flag)',
                    'CHH1517C': 'Children 15-17 years (flag)',
                    'INC_C': 'Income (grouped)',
                    'FAMINC_C': 'Family income (grouped)'
                }
                
                for var, label in filter_labels.items():
                    if var in all_filter_vars and all_filter_vars[var]:
                        selected_val = st.session_state.filters.get(var, None)
                        if selected_val is not None:
                            # Handle both single values and lists
                            if isinstance(selected_val, list):
                                if len(selected_val) > 0:
                                    selected_labels = []
                                    for val in selected_val:
                                        lbl = format_value(var, val)
                                        selected_labels.append(f"{lbl} ({val})")
                                    selected_display = "; ".join(selected_labels)
                                else:
                                    selected_display = "All"
                            else:
                                selected_label = format_value(var, selected_val)
                                selected_display = f"{selected_label} ({selected_val})"
                        else:
                            selected_display = "All"
                        
                        # Get all available options
                        options_list = []
                        for val in sorted(all_filter_vars[var]):
                            opt_label = format_value(var, val)
                            options_list.append(f"{opt_label} ({val})")
                        options_str = "; ".join(options_list[:10])  # Limit to first 10 for display
                        if len(options_list) > 10:
                            options_str += f"; ... ({len(options_list)} total options)"
                        
                        all_data.append([label, selected_display, options_str])
                
                all_data.append([""])
                all_data.append(["Number of Records Matching Criteria:", st.session_state.get('filtered_count', 'N/A')])
                all_data.append([""])
                all_data.append([""])
                
                # MIDDLE SECTION: Activity Categories
                all_data.append(["Activity Category Breakdown"])
                all_data.append(["Activity Category", "Mean Minutes Per Day", "Variance", "Standard Error", "Coefficient of Variation (%)"])
                
                if 'category_results' in st.session_state and st.session_state.category_results is not None:
                    category_results = st.session_state.category_results.copy()
                    for _, row in category_results.iterrows():
                        all_data.append([
                            row['Activity_Category'],
                            round(row['Mean_Minutes_Per_Day'], 2),
                            round(row['Variance'], 2),
                            round(row['Standard_Error'], 2),
                            round(row['Coefficient_of_Variation'], 2) if not pd.isna(row['Coefficient_of_Variation']) else ""
                        ])
                    
                    # Add total row
                    total_minutes = category_results['Mean_Minutes_Per_Day'].sum()
                    all_data.append(["TOTAL", round(total_minutes, 2), "", "", ""])
                    all_data.append(["Verification (should equal 1440):", round(total_minutes, 2)])
                
                all_data.append([""])
                all_data.append([""])
                
                # BOTTOM SECTION: Individual Activity Codes
                all_data.append(["Individual Activity Code Breakdown"])
                all_data.append(["Activity Code", "Activity Description", "Activity Category", 
                               "Mean Minutes Per Day", "Variance", "Standard Error", "Coefficient of Variation (%)"])
                
                display_cols = ['Activity_Code', 'Activity_Description', 'Activity_Category', 
                              'Mean_Minutes_Per_Day', 'Variance', 'Standard_Error', 'Coefficient_of_Variation']
                results_export = results_df[[c for c in display_cols if c in results_df.columns]].copy()
                
                for _, row in results_export.iterrows():
                    all_data.append([
                        row['Activity_Code'],
                        row['Activity_Description'],
                        row['Activity_Category'],
                        round(row['Mean_Minutes_Per_Day'], 2),
                        round(row['Variance'], 2),
                        round(row['Standard_Error'], 2),
                        round(row['Coefficient_of_Variation'], 2) if not pd.isna(row['Coefficient_of_Variation']) else ""
                    ])
                
                # Add total row
                total_minutes_detail = results_export['Mean_Minutes_Per_Day'].sum()
                all_data.append(["TOTAL", "", "", round(total_minutes_detail, 2), "", "", ""])
                all_data.append(["Verification (should equal 1440):", "", "", round(total_minutes_detail, 2), "", "", ""])
                
                # Convert to DataFrame and write
                export_df = pd.DataFrame(all_data)
                export_df.to_excel(writer, sheet_name='Time Estimates', index=False, header=False)
            
            # Format the Excel file
            output.seek(0)
            wb = load_workbook(output)
            ws = wb['Time Estimates']
            
            # Set print area and page setup
            max_row = ws.max_row
            max_col = ws.max_column
            ws.print_area = f'A1:{get_column_letter(max_col)}{max_row}'
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            # Style headers and important rows
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font = Font(bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format title row
            ws['A1'].font = title_font
            ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
            
            # Format section headers and highlight specific categories
            yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            in_category_section = False
            category_start_row = None
            category_header_row = None
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row), 1):
                cell_value = str(row[0].value) if row[0].value else ""
                
                # Track when we're in the Activity Category Breakdown section
                if "Activity Category Breakdown" in cell_value:
                    in_category_section = True
                    category_start_row = row_idx
                elif "Individual Activity Code Breakdown" in cell_value:
                    in_category_section = False
                
                # Check if this is the category header row (row after "Activity Category Breakdown")
                if in_category_section and row_idx == category_start_row + 1:
                    category_header_row = row_idx
                
                # Format section headers
                is_header = any(keyword in cell_value for keyword in ["Source:", "Filter Criteria:", "Activity Category Breakdown", 
                                                             "Individual Activity Code Breakdown", "TOTAL", "Verification"])
                if is_header:
                    for cell in row:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                # Highlight specific categories in Activity Category Breakdown section (skip header rows)
                if in_category_section and row_idx > category_start_row + 1 and row_idx != category_header_row:
                    first_cell_value = str(row[0].value) if row[0].value else ""
                    # Don't highlight if it's a header row (TOTAL, Verification)
                    if first_cell_value not in ["TOTAL", "Verification"] and first_cell_value in ["Household Work", "Shopping", "Childcare"]:
                        for cell in row:
                            cell.fill = yellow_fill
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # Save formatted workbook
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            excel_data = output.read()
            
            col_left, col_right = st.columns([1, 3])
            with col_left:
                st.download_button(
                    label="Download All Results (Excel)",
                    data=excel_data,
                    file_name="time_estimates_all.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="excel_all",
                    type="primary",
                    use_container_width=True
                )
        except ImportError:
            st.warning("Excel export requires openpyxl and Pillow. Install with: pip install openpyxl pillow")
        except Exception as e:
            st.error(f"Error creating Excel file: {e}")
            import traceback
            st.text(traceback.format_exc())

if __name__ == "__main__":
    main()

